from __future__ import annotations

import argparse
import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SOURCE_URL = (
    "https://brecke.inta.gatech.edu/wp-content/uploads/sites/19/2018/09/"
    "Conflict-Catalog-18-vars.xlsx"
)
SHEET_NAME = "Conflict Catalog 18 vars.xls"
EXPECTED_HEADERS = (
    "Common Name",
    "Name",
    "CountryCode",
    "NumberActors",
    "MilFatalities",
    "TotalFatalities",
    "StartYear",
    "StartMonth",
    "StartDay",
    "EndYear",
    "EndMonth",
    "EndDay",
    "Region",
    "Century",
    "Decade",
    "DurationD",
    "DurationM",
    "DurationY",
)
TRAILING_DATE_SUFFIX = re.compile(r",\s*\d{4}(-\d{2,4})?$")


def _missing(value: Any) -> bool:
    return value is None or value == -9 or str(value).strip() == "-9"


def _text(value: Any, *, question_is_missing: bool = False) -> str | None:
    if _missing(value):
        return None
    text = str(value).strip()
    if not text or (question_is_missing and text == "?"):
        return None
    return text


def _number(value: Any) -> int | float | None:
    if _missing(value):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else value
    text = str(value).strip()
    try:
        parsed = float(text)
    except ValueError:
        return None
    return int(parsed) if parsed.is_integer() else parsed


def strip_brecke_date_suffix(name: str) -> str:
    """Remove only the catalog's documented trailing four-digit date suffix."""

    return TRAILING_DATE_SUFFIX.sub("", name).strip()


def _interval_status(start_year: int | float | None, end_year: int | float | None) -> str:
    if start_year is None:
        return "missing_start"
    if end_year is None:
        return "open_end"
    if end_year < start_year:
        return "invalid_reversed"
    return "closed"


def parse_brecke_workbook(path: str | Path) -> list[dict[str, Any]]:
    source = Path(path)
    snapshot_sha256 = hashlib.sha256(source.read_bytes()).hexdigest()
    workbook = load_workbook(source, read_only=True, data_only=True)
    if workbook.sheetnames != [SHEET_NAME]:
        raise ValueError(
            f"Brecke workbook sheets changed: expected {[SHEET_NAME]}, "
            f"got {workbook.sheetnames}"
        )
    worksheet = workbook[SHEET_NAME]
    rows = worksheet.iter_rows(values_only=True)
    headers = tuple(next(rows))
    if headers != EXPECTED_HEADERS:
        raise ValueError(
            f"Brecke workbook headers changed: expected {EXPECTED_HEADERS}, got {headers}"
        )

    records: list[dict[str, Any]] = []
    for source_row, values in enumerate(rows, start=2):
        raw = dict(zip(EXPECTED_HEADERS, values, strict=True))
        name_raw = _text(raw["Name"])
        if name_raw is None:
            continue
        common_name = _text(raw["Common Name"], question_is_missing=True)
        war_name = strip_brecke_date_suffix(name_raw)
        aliases: list[str] = []
        seen_aliases: set[str] = set()
        for alias in (common_name, war_name):
            if alias and alias.casefold() not in seen_aliases:
                aliases.append(alias)
                seen_aliases.add(alias.casefold())

        start_year = _number(raw["StartYear"])
        end_year = _number(raw["EndYear"])
        region_code = _number(raw["Region"])
        record = {
            "aliases": aliases,
            "brecke_id": f"brecke-{source_row:04d}",
            "century_code": _number(raw["Century"]),
            "common_name": common_name,
            "country_code": _number(raw["CountryCode"]),
            "decade_code": _number(raw["Decade"]),
            "duration_days": _number(raw["DurationD"]),
            "duration_months": _number(raw["DurationM"]),
            "duration_years": _number(raw["DurationY"]),
            "end_day": _number(raw["EndDay"]),
            "end_month": _number(raw["EndMonth"]),
            "end_year": end_year,
            "interval_status": _interval_status(start_year, end_year),
            "military_fatalities": _number(raw["MilFatalities"]),
            "name_raw": name_raw,
            "number_actors": _number(raw["NumberActors"]),
            "outcome_available": False,
            "rating_use": "coverage_cross_check_only",
            "region_code": region_code,
            "region_label": (
                "Unclassified" if region_code is None else f"Brecke region {region_code}"
            ),
            "source": "brecke-conflict-catalog",
            "source_row": source_row,
            "source_snapshot": source.name,
            "source_snapshot_sha256": snapshot_sha256,
            "source_url": SOURCE_URL,
            "start_day": _number(raw["StartDay"]),
            "start_month": _number(raw["StartMonth"]),
            "start_year": start_year,
            "total_fatalities": _number(raw["TotalFatalities"]),
            "war_name": war_name,
        }
        records.append(record)
    workbook.close()
    return records


def write_registry(records: list[dict[str, Any]], destination: str | Path) -> None:
    output = Path(destination)
    output.parent.mkdir(parents=True, exist_ok=True)
    payload = "".join(
        json.dumps(record, ensure_ascii=False, sort_keys=True) + "\n"
        for record in records
    ).encode("utf-8")
    output.write_bytes(payload)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Build the Brecke war-name coverage registry from the local XLSX"
    )
    parser.add_argument(
        "--input",
        default=str(
            PROJECT_ROOT
            / "data/raw/brecke-conflict-catalog/Conflict-Catalog-18-vars.xlsx"
        ),
    )
    parser.add_argument(
        "--output",
        default=str(PROJECT_ROOT / "data/reference/brecke-wars.jsonl"),
    )
    args = parser.parse_args()
    records = parse_brecke_workbook(args.input)
    write_registry(records, args.output)
    print(f"wrote {len(records)} Brecke war-registry rows to {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
